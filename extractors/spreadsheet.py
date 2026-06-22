"""
Spreadsheet extractor for XLSX and CSV files.

XLSX  → one RawDocument per sheet (sheet name in metadata)
CSV   → one RawDocument for the whole file

Converts tabular data to plain text — the chunker and embedder
work on text, not dataframes.
"""

import csv
import hashlib
import io
import logging
from pathlib import Path

import aiofiles
import openpyxl

from app.core.exceptions import ExtractionError
from app.models.document import DocumentMetadata, RawDocument
from extractors.base import BaseExtractor

logger = logging.getLogger(__name__)


class SpreadsheetExtractor:
    """
    Extracts text from XLSX and CSV files.

    Converts tabular rows to pipe-delimited text:
        "Header1 | Header2 | Header3"
        "Value1  | Value2  | Value3"

    This format preserves column relationships while being
    embeddable as plain text by the embedding model.

    Satisfies BaseExtractor Protocol without explicit inheritance.
    """

    # [WHY] Pipe delimiter chosen over comma because CSV values
    # themselves often contain commas. Pipe is rarely in cell content.
    _COLUMN_DELIMITER: str = " | "
    _ROW_DELIMITER: str = "\n"

    def supported_mime_types(self) -> list[str]:
        """Return MIME types this extractor handles."""
        return [
            # [WHY] Two MIME types registered to one extractor instance.
            # ExtractorRegistry handles the mapping — no duplication.
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "text/csv",
        ]

    async def extract(
        self,
        file_path: Path,
        metadata: DocumentMetadata,
    ) -> list[RawDocument]:
        """
        Route to XLSX or CSV extraction based on MIME type.

        Args:
            file_path: Absolute path to the spreadsheet file.
            metadata:  Base DocumentMetadata to attach.

        Returns:
            List of RawDocument instances.

        Raises:
            ExtractionError: If the file cannot be read or parsed.
        """
        suffix = file_path.suffix.lower()

        # [WHY] Route on file suffix as secondary signal after MIME detection.
        # MIME type is detected by registry before calling extract().
        # Suffix routing here is a safety net for edge cases.
        if suffix in {".xlsx", ".xlsm"}:
            return await self._extract_xlsx(file_path, metadata)
        elif suffix == ".csv":
            return await self._extract_csv(file_path, metadata)
        else:
            raise ExtractionError(
                message=f"SpreadsheetExtractor received unexpected file: {file_path.name}",
                detail=f"Suffix '{suffix}' not handled",
            )

    async def _extract_xlsx(
        self,
        file_path: Path,
        metadata: DocumentMetadata,
    ) -> list[RawDocument]:
        """
        Extract one RawDocument per sheet from an XLSX file.

        Args:
            file_path: Path to the XLSX file.
            metadata:  Base metadata to copy per sheet.

        Returns:
            List of RawDocument — one per non-empty sheet.

        Raises:
            ExtractionError: If the workbook cannot be loaded.
        """
        # [WHY] Read file bytes async first, then parse sync.
        # openpyxl.load_workbook is synchronous. Reading the full
        # file async keeps I/O non-blocking then hands bytes
        # to openpyxl in memory — best of both worlds.
        try:
            async with aiofiles.open(file_path, "rb") as f:
                file_bytes = await f.read()
        except OSError as exc:
            raise ExtractionError(
                message=f"Cannot read XLSX file: {file_path.name}",
                detail=str(exc),
            ) from exc

        try:
            workbook = openpyxl.load_workbook(
                io.BytesIO(file_bytes),
                read_only=True,  # [WHY] read_only=True is faster and uses less memory
                data_only=True,  # [WHY] data_only=True returns cell values not formulas
            )
        except Exception as exc:
            raise ExtractionError(
                message=f"Cannot parse XLSX workbook: {file_path.name}",
                detail=str(exc),
            ) from exc

        raw_documents: list[RawDocument] = []

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows: list[str] = []

                for row in sheet.iter_rows(values_only=True):
                    # [WHY] Filter None cells but preserve empty string cells.
                    # None means the cell is truly empty.
                    # Empty string means someone explicitly cleared a cell.
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in cells):
                        rows.append(self._COLUMN_DELIMITER.join(cells))

                if not rows:
                    logger.warning(
                        "Skipping empty sheet",
                        extra={"file": file_path.name, "sheet": sheet_name},
                    )
                    continue

                content = self._ROW_DELIMITER.join(rows)
                content_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()

                sheet_metadata = metadata.model_copy(
                    update={
                        "content_hash": content_hash,
                        "metadata": {
                            **metadata.metadata,
                            "sheet_name": sheet_name,
                            "row_count": len(rows),
                        },
                    }
                )

                raw_documents.append(
                    RawDocument(metadata=sheet_metadata, content=content)
                )

        finally:
            workbook.close()

        if not raw_documents:
            raise ExtractionError(
                message=f"No extractable content in XLSX: {file_path.name}",
                detail="All sheets were empty",
            )

        logger.info(
            "XLSX extraction complete",
            extra={
                "file": file_path.name,
                "sheets_extracted": len(raw_documents),
            },
        )

        return raw_documents

    async def _extract_csv(
        self,
        file_path: Path,
        metadata: DocumentMetadata,
    ) -> list[RawDocument]:
        """
        Extract one RawDocument from a CSV file.

        Args:
            file_path: Path to the CSV file.
            metadata:  Base metadata to attach.

        Returns:
            Single-element list containing the extracted RawDocument.

        Raises:
            ExtractionError: If the file cannot be read or parsed.
        """
        try:
            async with aiofiles.open(file_path, encoding="utf-8", errors="replace") as f:
                raw_text = await f.read()
        except OSError as exc:
            raise ExtractionError(
                message=f"Cannot read CSV file: {file_path.name}",
                detail=str(exc),
            ) from exc

        try:
            reader = csv.reader(io.StringIO(raw_text))
            rows = [
                self._COLUMN_DELIMITER.join(cell.strip() for cell in row)
                for row in reader
                if any(cell.strip() for cell in row)
            ]
        except csv.Error as exc:
            raise ExtractionError(
                message=f"Cannot parse CSV: {file_path.name}",
                detail=str(exc),
            ) from exc

        if not rows:
            raise ExtractionError(
                message=f"No extractable content in CSV: {file_path.name}",
                detail="File was empty or contained only blank rows",
            )

        content = self._ROW_DELIMITER.join(rows)
        content_hash = hashlib.sha256(content.encode("utf-8")).hexdigest()

        csv_metadata = metadata.model_copy(
            update={
                "content_hash": content_hash,
                "metadata": {
                    **metadata.metadata,
                    "row_count": len(rows),
                },
            }
        )

        logger.info(
            "CSV extraction complete",
            extra={"file": file_path.name, "rows_extracted": len(rows)},
        )

        return [RawDocument(metadata=csv_metadata, content=content)]